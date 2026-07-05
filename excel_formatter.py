"""
excel_formatter.py
------------------
Tanggung jawab: Memformat tampilan file Excel setelah proses pengiriman selesai.
Menambahkan warna, emoji, dan style agar mudah dibaca operator.

Mengapa modul ini dibuat terpisah?
- Formatting adalah urusan "tampilan" — tidak ada hubungannya dengan
  logika bisnis (baca data, kirim email, simpan status).
- Prinsip Single Responsibility: excel_reader.py mengurusi DATA,
  excel_formatter.py mengurusi TAMPILAN.
- Jika suatu hari format berubah, cukup ubah file ini tanpa
  menyentuh logika pengiriman email sama sekali.
- Mudah dinonaktifkan jika tidak diperlukan (cukup tidak panggil format_excel).

Mengapa menggunakan openpyxl langsung, bukan pandas?
- pandas.to_excel() hanya menyimpan DATA — tidak bisa mengatur warna/style.
- openpyxl memberikan akses penuh ke setiap sel, baris, dan kolom.
- Warna yang diterapkan openpyxl tersimpan permanen di file .xlsx.

Alur kerja:
1. excel_reader.py menyimpan data via pandas (data sudah benar)
2. excel_formatter.py membuka file yang sama via openpyxl (untuk styling)
3. Formatter membaca kolom Status setiap baris
4. Terapkan warna sesuai status
5. Simpan kembali — file sekarang punya data + style

Hubungan dengan file lain:
- Dipanggil oleh main.py setelah save_excel() selesai.
- Membaca EXCEL_FILE dari config.py.
- Menggunakan logger dari logger.py.
- TIDAK mengubah data — hanya mengubah tampilan.
"""

from pathlib import Path
from typing import Dict, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_FILE
from logger import get_logger

logger = get_logger()


# =============================================================================
# DEFINISI WARNA
# =============================================================================
# Warna didefinisikan sebagai konstanta di satu tempat.
# Mengapa? Jika warna perlu diubah, cukup ubah di sini — tidak perlu
# mencari-cari di seluruh kode.
#
# Format warna openpyxl: HEX tanpa tanda # (contoh: "FF0000" = merah)

# Warna untuk sel STATUS (kolom Status saja)
COLOR_STATUS_SENT_BG     = "2E7D32"   # Hijau tua
COLOR_STATUS_SENT_FONT   = "FFFFFF"   # Putih
COLOR_STATUS_FAILED_BG   = "C62828"   # Merah tua
COLOR_STATUS_FAILED_FONT = "FFFFFF"   # Putih
COLOR_STATUS_SKIP_BG     = "F9A825"   # Kuning tua
COLOR_STATUS_SKIP_FONT   = "212121"   # Hitam

# Warna untuk seluruh BARIS (highlight lembut)
COLOR_ROW_SENT_BG   = "E8F5E9"   # Hijau sangat muda
COLOR_ROW_FAILED_BG = "FFEBEE"   # Merah sangat muda (merah muda)
COLOR_ROW_SKIP_BG   = "FFFDE7"   # Kuning sangat muda

# Warna header
COLOR_HEADER_BG   = "1565C0"   # Biru tua
COLOR_HEADER_FONT = "FFFFFF"   # Putih

# Warna baris bergantian (stripe) untuk baris yang tidak punya status khusus
COLOR_STRIPE_ODD  = "F5F5F5"   # Abu-abu sangat muda
COLOR_STRIPE_EVEN = "FFFFFF"   # Putih


# =============================================================================
# PEMETAAN STATUS → EMOJI & STYLE
# =============================================================================
# Semua aturan tampilan berdasarkan status dikumpulkan di satu dict.
# Ini membuat kode lebih mudah dibaca dan dikembangkan.
#
# Struktur: {
#   "nilai_status": {
#       "emoji"    : emoji yang ditambahkan ke teks status,
#       "cell_bg"  : warna background sel Status,
#       "cell_font": warna font sel Status,
#       "row_bg"   : warna background seluruh baris,
#   }
# }

STATUS_STYLE_MAP: Dict[str, Dict] = {
    "Terkirim": {
        "emoji"    : "✅",
        "cell_bg"  : COLOR_STATUS_SENT_BG,
        "cell_font": COLOR_STATUS_SENT_FONT,
        "row_bg"   : COLOR_ROW_SENT_BG,
    },
    "Gagal": {
        "emoji"    : "❌",
        "cell_bg"  : COLOR_STATUS_FAILED_BG,
        "cell_font": COLOR_STATUS_FAILED_FONT,
        "row_bg"   : COLOR_ROW_FAILED_BG,
    },
    "Dilewati": {
        "emoji"    : "⏭",
        "cell_bg"  : COLOR_STATUS_SKIP_BG,
        "cell_font": COLOR_STATUS_SKIP_FONT,
        "row_bg"   : COLOR_ROW_SKIP_BG,
    },
}


# =============================================================================
# FUNGSI UTAMA
# =============================================================================

def format_excel(filepath: Path = EXCEL_FILE) -> None:
    """
    Entry point utama — memformat seluruh file Excel.

    Urutan kerja:
    1. Buka file Excel yang sudah disimpan pandas (data sudah benar).
    2. Ambil sheet pertama (worksheet aktif).
    3. Format baris header.
    4. Sesuaikan lebar kolom otomatis.
    5. Format setiap baris data berdasarkan nilai kolom Status.
    6. Bekukan baris pertama (freeze panes) agar header selalu terlihat.
    7. Simpan file dengan style yang sudah diterapkan.

    Mengapa format dilakukan SETELAH pandas menyimpan data?
    Pandas tidak mendukung styling — dia hanya tahu data.
    Kita biarkan pandas melakukan tugasnya (simpan data dengan benar),
    lalu kita buka ulang filenya dengan openpyxl untuk menambah style.

    Args:
        filepath: Path ke file Excel yang akan diformat.
                  Default menggunakan EXCEL_FILE dari config.py.
    """
    if not filepath.exists():
        logger.warning(f"File Excel tidak ditemukan untuk diformat: {filepath}")
        return

    logger.info("Memformat tampilan Excel...")

    try:
        # Buka workbook yang sudah ada (data sudah tersimpan oleh pandas)
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active  # Ambil sheet pertama / sheet yang sedang aktif

        # Temukan nomor kolom Status secara dinamis
        # Mengapa dinamis? Jika urutan kolom berubah, kode tetap benar.
        status_col_idx = _find_column_index(ws, "Status")
        if status_col_idx is None:
            logger.warning("Kolom 'Status' tidak ditemukan — formatting dibatalkan.")
            wb.close()
            return

        # Jalankan setiap langkah formatting
        _format_header_row(ws)
        _auto_adjust_column_widths(ws)
        _format_data_rows(ws, status_col_idx)
        _freeze_header_row(ws)

        # Simpan file dengan style yang sudah diterapkan
        wb.save(filepath)
        wb.close()

        logger.info("Formatting Excel selesai — file siap dibuka operator.")

    except Exception as e:
        # Jangan biarkan error formatting menghentikan program
        # Data sudah tersimpan dengan benar — formatting gagal tidak kritis
        logger.warning(f"Formatting Excel gagal (data tetap tersimpan): {e}")


# =============================================================================
# FUNGSI FORMATTING DETAIL
# =============================================================================

def _format_header_row(ws: Worksheet) -> None:
    """
    Memformat baris pertama (header) dengan background biru dan font putih bold.

    Mengapa header perlu diformat?
    - Membantu operator langsung tahu mana judul kolom dan mana data.
    - Tampilan profesional dan konsisten.

    Args:
        ws: Worksheet yang akan diformat.
    """
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=COLOR_HEADER_BG
    )
    header_font = Font(
        color=COLOR_HEADER_FONT,
        bold=True,
        size=11,
    )
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=False,
    )

    # Iterasi semua sel di baris pertama (baris 1 = header)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_alignment

    # Tinggi baris header sedikit lebih besar agar nyaman dibaca
    ws.row_dimensions[1].height = 22


def _format_data_rows(ws: Worksheet, status_col_idx: int) -> None:
    """
    Memformat setiap baris data berdasarkan nilai kolom Status.

    Logika:
    - Baca nilai sel di kolom Status setiap baris.
    - Normalisasi nilai (hilangkan emoji lama jika ada).
    - Cari style yang sesuai di STATUS_STYLE_MAP.
    - Terapkan:
      a. Warna seluruh baris (highlight lembut).
      b. Warna + emoji + bold pada sel Status.

    Args:
        ws            : Worksheet yang akan diformat.
        status_col_idx: Nomor kolom (1-based) dari kolom Status.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Ambil sel Status di baris ini
        status_cell = ws.cell(row=row_idx, column=status_col_idx)
        raw_status  = str(status_cell.value or "").strip()

        # Bersihkan emoji lama jika program dijalankan ulang
        # (agar tidak menumpuk: "✅✅ Terkirim")
        clean_status = _strip_emoji_prefix(raw_status)

        # Cari style berdasarkan status bersih
        style = STATUS_STYLE_MAP.get(clean_status)

        if style:
            # --- Warnai seluruh baris dengan warna lembut ---
            row_fill = PatternFill(fill_type="solid", fgColor=style["row_bg"])
            for cell in row:
                cell.fill = row_fill
                # Rata kiri untuk semua data, tengah untuk status
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # --- Format sel Status secara khusus ---
            status_cell.value = f"{style['emoji']} {clean_status}"
            status_cell.fill  = PatternFill(
                fill_type="solid",
                fgColor=style["cell_bg"]
            )
            status_cell.font  = Font(
                color=style["cell_font"],
                bold=True,
                size=11,
            )
            status_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        else:
            # Status tidak dikenal atau kosong — beri warna stripe biasa
            stripe_color = COLOR_STRIPE_ODD if row_idx % 2 == 0 else COLOR_STRIPE_EVEN
            stripe_fill  = PatternFill(fill_type="solid", fgColor=stripe_color)
            for cell in row:
                cell.fill      = stripe_fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Tinggi baris yang konsisten
        ws.row_dimensions[row_idx].height = 18


def _auto_adjust_column_widths(ws: Worksheet) -> None:
    """
    Menyesuaikan lebar kolom secara otomatis berdasarkan konten terpanjang.

    Mengapa ini perlu?
    Pandas menyimpan data dengan lebar kolom default yang seringkali
    terlalu sempit sehingga teks terpotong. Fungsi ini menghitung
    panjang teks terpanjang di setiap kolom lalu mengatur lebarnya.

    Batas lebar:
    - Minimum: 10 karakter (agar tidak terlalu sempit)
    - Maksimum: 60 karakter (agar tidak terlalu lebar untuk kolom Error)

    Args:
        ws: Worksheet yang akan diformat.
    """
    MIN_WIDTH = 10
    MAX_WIDTH = 60

    for col_cells in ws.columns:
        # Hitung panjang konten terpanjang di kolom ini
        max_length = 0
        col_letter  = get_column_letter(col_cells[0].column)

        for cell in col_cells:
            if cell.value:
                # Ambil baris pertama saja jika ada newline (kolom Error bisa panjang)
                first_line = str(cell.value).split("\n")[0]
                max_length = max(max_length, len(first_line))

        # Tambah sedikit padding agar tidak terlalu mepet
        adjusted_width = max(MIN_WIDTH, min(max_length + 4, MAX_WIDTH))
        ws.column_dimensions[col_letter].width = adjusted_width


def _freeze_header_row(ws: Worksheet) -> None:
    """
    Membekukan baris pertama agar header selalu terlihat saat scroll ke bawah.

    Cara kerja freeze_panes:
    - ws.freeze_panes = "A2" artinya: bekukan semua baris di ATAS sel A2.
    - Hasilnya: baris 1 (header) selalu terlihat meskipun scroll ke baris 1000.

    Ini sangat berguna jika data Excel memiliki ratusan baris.

    Args:
        ws: Worksheet yang akan dibekukan.
    """
    ws.freeze_panes = "A2"


# =============================================================================
# FUNGSI HELPER INTERNAL
# =============================================================================

def _find_column_index(ws: Worksheet, column_name: str) -> Optional[int]:
    """
    Mencari nomor kolom (1-based) berdasarkan nama header.

    Mengapa tidak hardcode nomor kolom?
    Jika urutan kolom di Excel berubah, kode tetap bekerja dengan benar
    karena kita mencari berdasarkan nama, bukan posisi.

    Args:
        ws         : Worksheet yang akan dicari.
        column_name: Nama kolom yang dicari (case-sensitive).

    Returns:
        Optional[int]: Nomor kolom (1, 2, 3, ...) atau None jika tidak ditemukan.
    """
    for cell in ws[1]:  # Baris 1 adalah header
        if str(cell.value).strip() == column_name:
            return cell.column  # type: ignore[return-value]
    return None


def _strip_emoji_prefix(text: str) -> str:
    """
    Menghapus emoji dan spasi di awal teks status.

    Diperlukan agar saat program dijalankan ulang, emoji tidak menumpuk.
    Contoh: "✅ Terkirim" → "Terkirim"
            "❌ Gagal"    → "Gagal"
            "Terkirim"   → "Terkirim" (tidak berubah)

    Cara kerja:
    Iterasi karakter dari kiri, lewati semua karakter yang bukan
    huruf latin atau angka (emoji, spasi, dsb).

    Args:
        text: String status yang mungkin mengandung emoji.

    Returns:
        str: Status bersih tanpa emoji.
    """
    # Hapus spasi dulu
    text = text.strip()

    # Cari posisi huruf pertama (abjad atau angka)
    for i, char in enumerate(text):
        if char.isalpha() or char.isdigit():
            return text[i:].strip()

    return text
